"""Tahap 2 pipeline: cleaning, integration, and entity resolution."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.cell.cell import ERROR_CODES

from src.data_loader import find_raw_workbook, load_sheets
from src.entity_resolution import MatchThresholds, resolve_entities
from src.preprocessing import (
    clean_text_basic,
    normalize_place_name,
    normalize_url_spacing,
    parse_coordinate,
    parse_price,
    parse_rating,
    parse_review_date,
    standardize_columns,
    text_length,
    weak_sentiment_from_rating,
)


ROOT = Path(__file__).resolve().parents[1]
PROCESSED_DIR = ROOT / "data" / "processed"
INTERIM_DIR = ROOT / "data" / "interim"
REPORT_DIR = ROOT / "outputs" / "reports"
CONFIG_DIR = ROOT / "configs"
CONFIG_PATH = CONFIG_DIR / "config.yaml"

REVIEW_SPECS = {
    "wisata-v2": {
        "place_col": "place_name",
        "rating_col": "reviewer_rating",
        "review_col": "review_text",
        "published_col": "published_at",
        "collect_col": "collect_date",
        "category": "wisata",
    },
    "resto-hotel-v2": {
        "place_col": "place_name",
        "rating_col": "reviewer_rating",
        "review_col": "review_text",
        "published_col": "published_at",
        "collect_col": "collect_date",
        "category": "hotel_resto",
    },
    "tempat-wisata-v1": {
        "place_col": "place",
        "rating_col": "rating",
        "review_col": "review",
        "published_col": None,
        "collect_col": None,
        "category": "wisata",
    },
    "hotel-resto-v1": {
        "place_col": "place",
        "rating_col": "rating",
        "review_col": "review",
        "published_col": None,
        "collect_col": None,
        "category": "hotel_resto",
    },
}

METADATA_SPECS = {
    "wisata-metadata": {
        "place_col": "place_name",
        "category": "wisata",
        "price_col": "entry_fee",
        "coord_col": "lat_long",
        "rating_col": "place_rating",
        "address_col": "address",
        "type_col": "place_type",
        "status_col": "status",
        "facility_col": None,
    },
    "hotel-metadata": {
        "place_col": "place_name",
        "category": "hotel",
        "price_col": "price_per_head",
        "coord_col": "lat_long",
        "rating_col": "place_rating",
        "address_col": "address",
        "type_col": "place_type",
        "status_col": "status",
        "facility_col": "fasilitas",
    },
    "resto-metadata": {
        "place_col": "place_name",
        "category": "restoran",
        "price_col": "price_per_head",
        "coord_col": "lat_long",
        "rating_col": "place_rating",
        "address_col": "address",
        "type_col": "place_type",
        "status_col": "status",
        "facility_col": "fasilitas",
    },
}


def _value(row: pd.Series, column: str | None) -> Any:
    if not column or column not in row:
        return None
    return row[column]


def load_config(config_path: Path = CONFIG_PATH) -> dict[str, Any]:
    if not config_path.exists():
        return {}
    with config_path.open("r", encoding="utf-8") as file:
        return yaml.safe_load(file) or {}


def _clean_sheet(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = standardize_columns(df)
    cleaned = cleaned.dropna(how="all").copy()
    for column in cleaned.columns:
        if cleaned[column].dtype == object:
            cleaned[column] = cleaned[column].map(clean_text_basic)
    return cleaned


def _detect_unresolved_excel_cells(workbook_path: Path) -> pd.DataFrame:
    """Log formula/error cells from review-text columns that need manual handling."""
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    records = []
    for sheet_name, spec in REVIEW_SPECS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        header_map = {
            clean_text_basic(cell.value): idx + 1
            for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
        }
        review_original = spec["review_col"].replace("_", "-")
        review_col_idx = header_map.get(review_original) or header_map.get(spec["review_col"])
        if not review_col_idx:
            continue
        for row in ws.iter_rows(min_row=2):
            if len(row) < review_col_idx:
                continue
            cell = row[review_col_idx - 1]
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                records.append(
                    {
                        "sheet_name": sheet_name,
                        "cell": cell.coordinate,
                        "issue_type": "formula_in_review_text",
                        "raw_value": value,
                    }
                )
            elif cell.data_type == "e" or value in ERROR_CODES:
                records.append(
                    {
                        "sheet_name": sheet_name,
                        "cell": cell.coordinate,
                        "issue_type": "excel_error",
                        "raw_value": str(value),
                    }
                )
    return pd.DataFrame(records, columns=["sheet_name", "cell", "issue_type", "raw_value"])


def build_source_places(cleaned_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Collect place names from metadata and review sheets for entity resolution."""
    records = []
    for sheet_name, spec in METADATA_SPECS.items():
        df = cleaned_sheets.get(sheet_name)
        if df is None:
            continue
        for _, row in df.iterrows():
            place_name = clean_text_basic(_value(row, spec["place_col"]))
            if place_name is None:
                continue
            coord = parse_coordinate(_value(row, spec["coord_col"]))
            records.append(
                {
                    "source_sheet": sheet_name,
                    "source_priority": 0,
                    "source_place_name": place_name,
                    "normalized_place_name": normalize_place_name(place_name),
                    "place_category": spec["category"],
                    "place_type": clean_text_basic(_value(row, spec["type_col"])),
                    "address": clean_text_basic(_value(row, spec["address_col"])),
                    "latitude": coord["latitude"],
                    "longitude": coord["longitude"],
                }
            )

    for sheet_name, spec in REVIEW_SPECS.items():
        df = cleaned_sheets.get(sheet_name)
        if df is None:
            continue
        for place_name in df[spec["place_col"]].dropna().map(clean_text_basic).dropna().unique():
            records.append(
                {
                    "source_sheet": sheet_name,
                    "source_priority": 1,
                    "source_place_name": place_name,
                    "normalized_place_name": normalize_place_name(place_name),
                    "place_category": spec["category"],
                    "place_type": None,
                    "address": None,
                    "latitude": None,
                    "longitude": None,
                }
            )
    return pd.DataFrame(records)


def enrich_places_master(places: pd.DataFrame, cleaned_sheets: dict[str, pd.DataFrame], entity_mapping: pd.DataFrame) -> pd.DataFrame:
    metadata_rows = []
    for sheet_name, spec in METADATA_SPECS.items():
        df = cleaned_sheets.get(sheet_name)
        if df is None:
            continue
        for _, row in df.iterrows():
            place_name = clean_text_basic(_value(row, spec["place_col"]))
            if place_name is None:
                continue
            normalized = normalize_place_name(place_name)
            mapping = entity_mapping[
                (entity_mapping["source_sheet"] == sheet_name)
                & (entity_mapping["normalized_place_name"] == normalized)
            ]
            if mapping.empty:
                continue
            coord = parse_coordinate(_value(row, spec["coord_col"]))
            price = parse_price(_value(row, spec["price_col"]))
            metadata_rows.append(
                {
                    "canonical_place_id": mapping.iloc[0]["canonical_place_id"],
                    "metadata_source_sheet": sheet_name,
                    "metadata_place_name": place_name,
                    "metadata_place_category": spec["category"],
                    "metadata_place_type": clean_text_basic(_value(row, spec["type_col"])),
                    "place_rating": parse_rating(_value(row, spec["rating_col"])),
                    "place_rating_original": _value(row, spec["rating_col"]),
                    "address": clean_text_basic(_value(row, spec["address_col"])),
                    "status": clean_text_basic(_value(row, spec["status_col"])),
                    "facility_text": clean_text_basic(_value(row, spec["facility_col"])),
                    "latitude": coord["latitude"],
                    "longitude": coord["longitude"],
                    "coordinate_original": coord["original"],
                    "coordinate_parsing_success": coord["parsing_success"],
                    "min_price": price["min_price"],
                    "max_price": price["max_price"],
                    "is_free": price["is_free"],
                    "price_text_original": price["price_text_original"],
                    "price_parsing_success": price["parsing_success"],
                    "price_is_missing": price["is_missing"],
                    "price_is_inferred": price["is_inferred"],
                }
            )

    metadata = pd.DataFrame(metadata_rows)
    if metadata.empty:
        return places

    merged = places.merge(
        metadata.sort_values("metadata_source_sheet").drop_duplicates("canonical_place_id"),
        on="canonical_place_id",
        how="left",
        suffixes=("", "_metadata"),
    )
    merged["place_category"] = merged["metadata_place_category"].combine_first(merged["place_category"])
    merged["place_type"] = merged["metadata_place_type"].combine_first(merged["place_type"])
    merged["address"] = merged["address_metadata"].combine_first(merged["address"])
    merged["latitude"] = merged["latitude_metadata"].combine_first(merged["latitude"])
    merged["longitude"] = merged["longitude_metadata"].combine_first(merged["longitude"])
    drop_cols = [col for col in merged.columns if col.endswith("_metadata")]
    merged = merged.drop(columns=drop_cols)
    return merged


def build_reviews(cleaned_sheets: dict[str, pd.DataFrame], entity_mapping: pd.DataFrame, places: pd.DataFrame) -> pd.DataFrame:
    place_lookup = places.set_index("canonical_place_id")
    mapping_lookup = entity_mapping.drop_duplicates(["source_sheet", "normalized_place_name"]).set_index(
        ["source_sheet", "normalized_place_name"]
    )
    records = []
    for sheet_name, spec in REVIEW_SPECS.items():
        df = cleaned_sheets.get(sheet_name)
        if df is None:
            continue
        for idx, row in df.iterrows():
            place_name = clean_text_basic(_value(row, spec["place_col"]))
            if place_name is None:
                continue
            normalized = normalize_place_name(place_name)
            canonical_place_id = None
            if (sheet_name, normalized) in mapping_lookup.index:
                canonical_place_id = mapping_lookup.loc[(sheet_name, normalized)]["canonical_place_id"]
            place_info = place_lookup.loc[canonical_place_id] if canonical_place_id in place_lookup.index else {}
            rating = parse_rating(_value(row, spec["rating_col"]))
            raw_review = clean_text_basic(_value(row, spec["review_col"]))
            clean_review = normalize_url_spacing(raw_review)
            parsed_date = parse_review_date(_value(row, spec["published_col"]), _value(row, spec["collect_col"]))
            records.append(
                {
                    "review_id": f"{sheet_name}_{idx + 2:06d}",
                    "canonical_place_id": canonical_place_id,
                    "place_name": place_name,
                    "place_category": place_info.get("place_category", spec["category"]) if isinstance(place_info, pd.Series) else spec["category"],
                    "reviewer_rating": rating,
                    "review_text_raw": raw_review,
                    "review_text_clean": clean_review,
                    "review_date": parsed_date["review_date"],
                    "review_date_parsing_success": parsed_date["parsing_success"],
                    "review_date_is_inferred": parsed_date["is_inferred"],
                    "review_date_source": parsed_date["source"],
                    "source_sheet": sheet_name,
                    "latitude": place_info.get("latitude") if isinstance(place_info, pd.Series) else None,
                    "longitude": place_info.get("longitude") if isinstance(place_info, pd.Series) else None,
                    "is_duplicate": False,
                    "text_length": text_length(clean_review),
                    "weak_sentiment_label": weak_sentiment_from_rating(rating),
                }
            )
    reviews = pd.DataFrame(records)
    if reviews.empty:
        return reviews
    duplicate_subset = ["canonical_place_id", "reviewer_rating", "review_text_clean", "review_date"]
    reviews["is_duplicate"] = reviews.duplicated(subset=duplicate_subset, keep="first")
    return reviews


def run_cleaning() -> dict[str, Any]:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    INTERIM_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    workbook_path = find_raw_workbook()
    raw_sheets = load_sheets(workbook_path)
    cleaned_sheets = {sheet_name: _clean_sheet(df) for sheet_name, df in raw_sheets.items()}

    unresolved = _detect_unresolved_excel_cells(workbook_path)
    unresolved.to_csv(REPORT_DIR / "unresolved_excel_cells.csv", index=False)

    source_places = build_source_places(cleaned_sheets)
    source_places.to_csv(INTERIM_DIR / "source_places.csv", index=False)

    config = load_config()
    entity_config = config.get("entity_resolution", {})
    thresholds = MatchThresholds(
        fuzzy_auto=float(entity_config.get("fuzzy_auto", 94.0)),
        fuzzy_review=float(entity_config.get("fuzzy_review", 88.0)),
        tfidf_auto=float(entity_config.get("tfidf_auto", 0.94)),
        tfidf_review=float(entity_config.get("tfidf_review", 0.86)),
    )
    places, entity_mapping = resolve_entities(source_places, thresholds=thresholds)
    places = enrich_places_master(places, cleaned_sheets, entity_mapping)
    reviews = build_reviews(cleaned_sheets, entity_mapping, places)

    places.to_parquet(PROCESSED_DIR / "places_master.parquet", index=False)
    entity_mapping.to_parquet(PROCESSED_DIR / "entity_mapping.parquet", index=False)
    reviews.to_parquet(PROCESSED_DIR / "reviews_clean.parquet", index=False)

    review_candidates = entity_mapping[entity_mapping["needs_manual_review"]].copy()
    review_candidates.to_csv(REPORT_DIR / "entity_matches_for_review.csv", index=False)

    summary = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "workbook_path": str(workbook_path),
        "cleaned_sheet_count": len(cleaned_sheets),
        "source_place_rows": int(len(source_places)),
        "canonical_place_count": int(len(places)),
        "entity_mapping_rows": int(len(entity_mapping)),
        "entity_matches_needing_manual_review": int(len(review_candidates)),
        "review_rows": int(len(reviews)),
        "review_rows_with_text": int((reviews["text_length"] > 0).sum()) if not reviews.empty else 0,
        "duplicate_review_rows": int(reviews["is_duplicate"].sum()) if not reviews.empty else 0,
        "unresolved_excel_cells": int(len(unresolved)),
        "outputs": {
            "places_master": str(PROCESSED_DIR / "places_master.parquet"),
            "entity_mapping": str(PROCESSED_DIR / "entity_mapping.parquet"),
            "reviews_clean": str(PROCESSED_DIR / "reviews_clean.parquet"),
            "entity_matches_for_review": str(REPORT_DIR / "entity_matches_for_review.csv"),
            "unresolved_excel_cells": str(REPORT_DIR / "unresolved_excel_cells.csv"),
        },
    }
    (REPORT_DIR / "cleaning_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Run TobaPulse cleaning and entity resolution.")
    parser.parse_args()
    summary = run_cleaning()
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
