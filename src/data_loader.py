"""Workbook loading and profiling utilities for TobaPulse.

This module is intentionally reusable by notebooks and scripts. It reads the raw
Excel workbook, profiles every sheet, detects formula/error cells with openpyxl,
and writes exploration reports plus simple matplotlib figures.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
os.environ.setdefault("MPLCONFIGDIR", str(ROOT / "outputs" / ".matplotlib"))
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import ERROR_CODES

from src.preprocessing import (
    clean_text_basic,
    normalize_column_name,
    normalize_place_name,
    parse_rating,
    text_length,
)


RAW_DIR = ROOT / "data" / "raw"
REPORT_DIR = ROOT / "outputs" / "reports"
FIGURE_DIR = ROOT / "outputs" / "figures"

PLACE_COLUMN_CANDIDATES = {
    "place",
    "place_name",
    "nama_atraksi",
    "objek_destinasi_wisata",
    "transport_name",
    "kuliner_name",
}
REVIEW_COLUMN_CANDIDATES = {"review", "review_text", "rangkuman_paragraf_awal", "detail", "description"}
RATING_COLUMN_CANDIDATES = {"rating", "reviewer_rating", "place_rating", "rating_bpodt"}
COORD_COLUMN_CANDIDATES = {"lat_long", "latitude", "longitude", "koordinat", "coordinates"}
PRICE_COLUMN_CANDIDATES = {"entry_fee", "price", "price_per_head", "htm"}
DATE_COLUMN_CANDIDATES = {"published_at", "collect_date", "date", "tanggal"}


def find_raw_workbook(raw_dir: Path = RAW_DIR) -> Path:
    """Return the first xlsx file from data/raw, falling back to project root."""
    candidates = sorted(raw_dir.glob("*.xlsx")) if raw_dir.exists() else []
    if not candidates:
        candidates = sorted(ROOT.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No .xlsx workbook found in data/raw or project root.")
    return candidates[0]


def load_sheets(workbook_path: Path | None = None) -> dict[str, pd.DataFrame]:
    """Load all workbook sheets as object DataFrames with original column names."""
    path = workbook_path or find_raw_workbook()
    return pd.read_excel(path, sheet_name=None, dtype=object, engine="openpyxl")


def normalized_columns(df: pd.DataFrame) -> dict[str, str]:
    """Map original column names to normalized snake_case names."""
    return {str(column): normalize_column_name(column) for column in df.columns}


def _column_by_candidates(df: pd.DataFrame, candidates: set[str]) -> str | None:
    mapping = normalized_columns(df)
    for original, normalized in mapping.items():
        if normalized in candidates:
            return original
    return None


def _columns_by_candidates(df: pd.DataFrame, candidates: set[str]) -> list[str]:
    mapping = normalized_columns(df)
    return [original for original, normalized in mapping.items() if normalized in candidates]


def inspect_excel_cells(workbook_path: Path) -> dict[str, dict[str, Any]]:
    """Detect formula and Excel error cells using openpyxl with data_only=False."""
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    results: dict[str, dict[str, Any]] = {}
    for ws in workbook.worksheets:
        formulas: list[dict[str, str]] = []
        errors: list[dict[str, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": value})
                if cell.data_type == "e" or value in ERROR_CODES:
                    errors.append({"cell": cell.coordinate, "value": str(value)})
        results[ws.title] = {
            "formula_cell_count": len(formulas),
            "formula_cell_examples": formulas[:25],
            "error_cell_count": len(errors),
            "error_cell_examples": errors[:25],
        }
    return results


def _infer_format_examples(series: pd.Series, max_examples: int = 10) -> list[str]:
    values = []
    for value in series.dropna():
        cleaned = clean_text_basic(value)
        if cleaned and cleaned not in values:
            values.append(cleaned)
        if len(values) >= max_examples:
            break
    return values


def _rating_distribution(series: pd.Series) -> dict[str, int]:
    parsed = series.map(parse_rating).dropna()
    rounded = parsed.round(1).astype(str)
    return {key: int(value) for key, value in Counter(rounded).items()}


def _top_missing(missing: dict[str, int], limit: int = 10) -> dict[str, int]:
    return dict(sorted(missing.items(), key=lambda item: item[1], reverse=True)[:limit])


def profile_sheet(sheet_name: str, df: pd.DataFrame, excel_cells: dict[str, Any]) -> dict[str, Any]:
    """Build a detailed profile for one sheet."""
    col_map = normalized_columns(df)
    place_col = _column_by_candidates(df, PLACE_COLUMN_CANDIDATES)
    review_col = _column_by_candidates(df, REVIEW_COLUMN_CANDIDATES)
    rating_col = _column_by_candidates(df, RATING_COLUMN_CANDIDATES)
    coord_cols = _columns_by_candidates(df, COORD_COLUMN_CANDIDATES)
    price_cols = _columns_by_candidates(df, PRICE_COLUMN_CANDIDATES)
    date_cols = _columns_by_candidates(df, DATE_COLUMN_CANDIDATES)

    missing = {str(col): int(df[col].isna().sum()) for col in df.columns}
    dtypes = {str(col): str(dtype) for col, dtype in df.dtypes.items()}
    sample_values = {
        str(col): [str(value) for value in df[col].dropna().head(5).tolist()]
        for col in df.columns
    }

    unique_places = 0
    top_places: dict[str, int] = {}
    normalized_place_examples: list[dict[str, str]] = []
    if place_col:
        place_series = df[place_col].dropna().map(clean_text_basic).dropna()
        unique_places = int(place_series.nunique())
        top_places = {str(k): int(v) for k, v in place_series.value_counts().head(15).items()}
        place_pairs = []
        for raw_name in place_series.head(100):
            normalized = normalize_place_name(raw_name)
            if normalized and normalized != raw_name.lower():
                place_pairs.append({"raw": raw_name, "normalized": normalized})
            if len(place_pairs) >= 10:
                break
        normalized_place_examples = place_pairs

    review_stats: dict[str, Any] = {}
    if review_col:
        lengths = df[review_col].map(text_length)
        review_stats = {
            "review_column": str(review_col),
            "reviews_with_text": int((lengths > 0).sum()),
            "reviews_without_text": int((lengths == 0).sum()),
            "text_length_min": int(lengths.min()) if len(lengths) else 0,
            "text_length_median": float(lengths.median()) if len(lengths) else 0.0,
            "text_length_mean": float(lengths.mean()) if len(lengths) else 0.0,
            "text_length_max": int(lengths.max()) if len(lengths) else 0,
        }

    rating_stats: dict[str, Any] = {}
    if rating_col:
        parsed = df[rating_col].map(parse_rating)
        rating_stats = {
            "rating_column": str(rating_col),
            "valid_rating_count": int(parsed.notna().sum()),
            "invalid_or_missing_rating_count": int(parsed.isna().sum()),
            "min_rating": float(parsed.min()) if parsed.notna().any() else None,
            "max_rating": float(parsed.max()) if parsed.notna().any() else None,
            "mean_rating": float(parsed.mean()) if parsed.notna().any() else None,
            "distribution": _rating_distribution(df[rating_col]),
        }

    return {
        "sheet_name": sheet_name,
        "rows": int(df.shape[0]),
        "columns": int(df.shape[1]),
        "column_names": [str(col) for col in df.columns],
        "normalized_column_names": col_map,
        "dtypes": dtypes,
        "missing_values": missing,
        "top_missing_values": _top_missing(missing),
        "duplicate_rows": int(df.duplicated().sum()),
        "place_column": str(place_col) if place_col else None,
        "unique_places": unique_places,
        "top_places": top_places,
        "review_stats": review_stats,
        "rating_stats": rating_stats,
        "coordinate_columns": [str(col) for col in coord_cols],
        "coordinate_examples": {str(col): _infer_format_examples(df[col]) for col in coord_cols},
        "price_columns": [str(col) for col in price_cols],
        "price_examples": {str(col): _infer_format_examples(df[col]) for col in price_cols},
        "date_columns": [str(col) for col in date_cols],
        "date_examples": {str(col): _infer_format_examples(df[col]) for col in date_cols},
        "inconsistent_place_examples": normalized_place_examples,
        **excel_cells,
        "sample_values": sample_values,
    }


def compare_place_names(sheets: dict[str, pd.DataFrame]) -> dict[str, Any]:
    """Compare normalized place-name sets across sheets."""
    per_sheet: dict[str, set[str]] = {}
    raw_examples: dict[str, dict[str, str]] = {}
    for sheet_name, df in sheets.items():
        place_col = _column_by_candidates(df, PLACE_COLUMN_CANDIDATES)
        if not place_col:
            continue
        names = {}
        for value in df[place_col].dropna():
            cleaned = clean_text_basic(value)
            normalized = normalize_place_name(value)
            if cleaned and normalized:
                names.setdefault(normalized, cleaned)
        per_sheet[sheet_name] = set(names)
        raw_examples[sheet_name] = names

    all_names = set().union(*per_sheet.values()) if per_sheet else set()
    overlap_rows = []
    for sheet_name, names in per_sheet.items():
        only_here = sorted(names - set().union(*(v for k, v in per_sheet.items() if k != sheet_name)))
        overlap_rows.append(
            {
                "sheet_name": sheet_name,
                "unique_normalized_places": len(names),
                "places_not_seen_in_other_sheets_count": len(only_here),
                "places_not_seen_in_other_sheets_examples": [
                    raw_examples[sheet_name][name] for name in only_here[:10]
                ],
            }
        )
    return {
        "total_unique_normalized_place_names": len(all_names),
        "by_sheet": overlap_rows,
    }


def metadata_quality(profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Summarize metadata completeness for metadata-like sheets."""
    rows = []
    for profile in profiles:
        sheet = profile["sheet_name"]
        if "metadata" not in sheet.lower() and sheet not in {"transportasi", "kuliner", "waktu operasional destinasi"}:
            continue
        total_cells = profile["rows"] * profile["columns"]
        missing_cells = sum(profile["missing_values"].values())
        completeness = 1.0 - (missing_cells / total_cells) if total_cells else 0.0
        rows.append(
            {
                "sheet_name": sheet,
                "rows": profile["rows"],
                "columns": profile["columns"],
                "missing_cells": int(missing_cells),
                "total_cells": int(total_cells),
                "completeness_ratio": round(completeness, 4),
                "place_column": profile["place_column"],
                "unique_places": profile["unique_places"],
            }
        )
    return rows


def profile_workbook(workbook_path: Path | None = None) -> dict[str, Any]:
    """Profile every sheet in the raw workbook and return a JSON-serializable dict."""
    path = workbook_path or find_raw_workbook()
    sheets = load_sheets(path)
    cell_report = inspect_excel_cells(path)
    profiles = [
        profile_sheet(name, df, cell_report.get(name, {}))
        for name, df in sheets.items()
    ]
    return {
        "generated_at": datetime.now().astimezone().isoformat(),
        "workbook_path": str(path),
        "file_name": path.name,
        "sheet_count": len(sheets),
        "sheets": profiles,
        "place_name_comparison": compare_place_names(sheets),
        "metadata_quality": metadata_quality(profiles),
    }


def write_profile(profile: dict[str, Any], report_dir: Path = REPORT_DIR) -> None:
    """Write JSON profile and CSV quality summary."""
    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / "data_profile.json").write_text(
        json.dumps(profile, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    rows = []
    for sheet in profile["sheets"]:
        rows.append(
            {
                "sheet_name": sheet["sheet_name"],
                "rows": sheet["rows"],
                "columns": sheet["columns"],
                "duplicate_rows": sheet["duplicate_rows"],
                "missing_cells": sum(sheet["missing_values"].values()),
                "formula_cell_count": sheet["formula_cell_count"],
                "error_cell_count": sheet["error_cell_count"],
                "place_column": sheet["place_column"],
                "unique_places": sheet["unique_places"],
                "review_column": sheet.get("review_stats", {}).get("review_column"),
                "reviews_with_text": sheet.get("review_stats", {}).get("reviews_with_text"),
                "reviews_without_text": sheet.get("review_stats", {}).get("reviews_without_text"),
                "rating_column": sheet.get("rating_stats", {}).get("rating_column"),
                "valid_rating_count": sheet.get("rating_stats", {}).get("valid_rating_count"),
                "invalid_or_missing_rating_count": sheet.get("rating_stats", {}).get("invalid_or_missing_rating_count"),
            }
        )
    pd.DataFrame(rows).to_csv(report_dir / "data_quality_summary.csv", index=False)


def _save_bar(series: pd.Series, title: str, xlabel: str, ylabel: str, path: Path, rotate: bool = False) -> None:
    plt.figure(figsize=(10, 6))
    series.plot(kind="bar")
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    if rotate:
        plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(path, dpi=150)
    plt.close()


def generate_figures(profile: dict[str, Any], workbook_path: Path | None = None, figure_dir: Path = FIGURE_DIR) -> None:
    """Generate required exploration figures with matplotlib."""
    figure_dir.mkdir(parents=True, exist_ok=True)
    path = workbook_path or Path(profile["workbook_path"])
    sheets = load_sheets(path)

    rating_values = []
    review_lengths = []
    review_counts = {}
    top_place_counts = Counter()
    missing_ratios = {}
    published_dates = []

    for sheet_name, df in sheets.items():
        missing_ratios[sheet_name] = float(df.isna().sum().sum() / (df.shape[0] * df.shape[1])) if df.shape[0] and df.shape[1] else 0.0
        review_col = _column_by_candidates(df, REVIEW_COLUMN_CANDIDATES)
        rating_col = _column_by_candidates(df, RATING_COLUMN_CANDIDATES)
        place_col = _column_by_candidates(df, PLACE_COLUMN_CANDIDATES)
        date_cols = _columns_by_candidates(df, DATE_COLUMN_CANDIDATES)

        if review_col:
            lengths = df[review_col].map(text_length)
            review_lengths.extend(lengths[lengths > 0].tolist())
            review_counts[sheet_name] = int((lengths > 0).sum())
        if rating_col:
            parsed = df[rating_col].map(parse_rating).dropna()
            rating_values.extend(parsed.tolist())
        if place_col:
            for place, count in df[place_col].dropna().map(clean_text_basic).dropna().value_counts().head(30).items():
                top_place_counts[str(place)] += int(count)
        for date_col in date_cols:
            parsed_dates = pd.to_datetime(df[date_col], errors="coerce")
            published_dates.extend(parsed_dates.dropna().tolist())

    if rating_values:
        rating_series = pd.Series(rating_values).round(1).value_counts().sort_index()
        _save_bar(rating_series, "Distribusi Rating", "Rating", "Jumlah", figure_dir / "rating_distribution.png")

    if missing_ratios:
        missing_series = pd.Series(missing_ratios).sort_values(ascending=False)
        _save_bar(missing_series, "Rasio Missing Value per Sheet", "Sheet", "Rasio Missing", figure_dir / "missing_value_by_sheet.png", rotate=True)

    if review_counts:
        review_series = pd.Series(review_counts).sort_values(ascending=False)
        _save_bar(review_series, "Jumlah Ulasan dengan Teks per Sheet", "Sheet", "Jumlah Ulasan", figure_dir / "review_count_by_category.png", rotate=True)

    if review_lengths:
        plt.figure(figsize=(10, 6))
        pd.Series(review_lengths).clip(upper=1000).plot(kind="hist", bins=40)
        plt.title("Distribusi Panjang Teks Ulasan")
        plt.xlabel("Panjang Teks")
        plt.ylabel("Jumlah Ulasan")
        plt.tight_layout()
        plt.savefig(figure_dir / "review_text_length_distribution.png", dpi=150)
        plt.close()

    if top_place_counts:
        top_places = pd.Series(dict(top_place_counts.most_common(20))).sort_values()
        plt.figure(figsize=(10, 8))
        top_places.plot(kind="barh")
        plt.title("Tempat dengan Ulasan Terbanyak")
        plt.xlabel("Jumlah Baris/Ulasan")
        plt.ylabel("Tempat")
        plt.tight_layout()
        plt.savefig(figure_dir / "top_places_by_review_count.png", dpi=150)
        plt.close()

    if published_dates:
        date_series = pd.Series(published_dates).dt.to_period("M").astype(str).value_counts().sort_index()
        if not date_series.empty:
            _save_bar(date_series.tail(36), "Distribusi Waktu Publikasi Ulasan", "Bulan", "Jumlah", figure_dir / "review_time_distribution.png", rotate=True)


def run_profile() -> dict[str, Any]:
    workbook_path = find_raw_workbook()
    profile = profile_workbook(workbook_path)
    write_profile(profile)
    generate_figures(profile, workbook_path)
    return profile


def main() -> None:
    parser = argparse.ArgumentParser(description="Profile the raw TobaPulse workbook.")
    parser.add_argument("--workbook", type=Path, default=None, help="Optional workbook path.")
    args = parser.parse_args()
    profile = profile_workbook(args.workbook)
    write_profile(profile)
    generate_figures(profile, args.workbook)
    print(json.dumps({"sheet_count": profile["sheet_count"], "reports_dir": str(REPORT_DIR), "figures_dir": str(FIGURE_DIR)}, indent=2))


if __name__ == "__main__":
    main()
